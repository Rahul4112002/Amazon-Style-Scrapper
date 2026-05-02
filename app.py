"""
Amazon ASIN Scraper — FastAPI + httpx + BeautifulSoup
======================================================
Upload an Excel sheet with Amazon ASINs, scrape each product page,
stream real-time logs via SSE, and export results to Excel.

⚠  Educational / personal-use only. Scraping may violate Amazon's ToS.
"""

from __future__ import annotations

import asyncio
import base64
import json
import logging
import os
import random
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import httpx
import pandas as pd
from bs4 import BeautifulSoup
from fastapi import FastAPI, File, Request, UploadFile
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------
app = FastAPI(title="Amazon ASIN Scraper")
templates = Jinja2Templates(directory="templates")

IS_VERCEL = os.environ.get("VERCEL") == "1"
STORAGE_ROOT = Path("/tmp") if IS_VERCEL else Path(".")

DOWNLOADS_DIR = STORAGE_ROOT / "downloads"
DOWNLOADS_DIR.mkdir(exist_ok=True)

UPLOADS_DIR = STORAGE_ROOT / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("asin_scraper")

# ---------------------------------------------------------------------------
# Global scraper state
# ---------------------------------------------------------------------------
scraper_state: dict[str, Any] = {
    "running": False,
    "stop_event": asyncio.Event(),
    "logs": [],          # list[str]
    "products": [],      # list[dict]
    "total_items": 0,
    "scraped_count": 0,
    "failed_count": 0,
    "captcha_blocked": [],  # ASINs blocked by CAPTCHA for retry
    "output_file": None,
    "task": None,
}

# Rotate user agents to avoid detection
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
]

def get_headers() -> dict:
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Cache-Control": "max-age=0",
    }

MAX_CONCURRENCY = 3  # Lower concurrency for Amazon to avoid rate limiting

DETAIL_LABEL_TO_COLUMN = {
    "Product Dimensions": "product_dimensions",
    "Date First Available": "date_first_available",
    "Manufacturer": "manufacturer",
    "Item model number": "item_model_number",
    "Country of Origin": "country_of_origin",
    "Department": "department",
    "Packer": "packer",
    "Importer": "importer",
    "Item Weight": "item_weight",
    "Item Dimensions LxWxH": "item_dimensions_lxwxh",
    "Net Quantity": "net_quantity",
    "Generic Name": "generic_name",
    "Best Sellers Rank": "best_sellers_rank",
    # Keep ASIN in the primary asin column instead of creating a duplicate.
    "ASIN": "asin",
}

DETAIL_COLUMNS = [
    "product_dimensions",
    "date_first_available",
    "manufacturer",
    "item_model_number",
    "country_of_origin",
    "department",
    "packer",
    "importer",
    "item_weight",
    "item_dimensions_lxwxh",
    "net_quantity",
    "generic_name",
    "best_sellers_rank",
]

BASE_EXPORT_COLUMNS = [
    "asin",
    "url",
    "title",
    "brand",
    "current_price",
    "original_price",
    "discount",
    "rating",
    "rating_count",
    "review_count",
    "description",
    "category",
    "images",
    "video_url",
    "colors",
    "sizes",
    "material",
    "seller",
    "availability",
    "features",
]

EXPORT_COLUMNS = BASE_EXPORT_COLUMNS + DETAIL_COLUMNS + ["product_details"]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _log(msg: str) -> None:
    line = f"{_ts()} | {msg}"
    scraper_state["logs"].append(line)
    logger.info(msg)


def _clean_text(text: str) -> str:
    """Clean and normalize text."""
    if not text:
        return ""
    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _clean_detail_text(text: Any) -> str:
    """Normalize Amazon detail text, including hidden direction marks."""
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    text = str(text)
    text = re.sub(r"[\u200e\u200f\u202a-\u202e]", "", text)
    text = text.replace("\xa0", " ")
    return _clean_text(text)


def _normalize_detail_label(label: Any) -> str:
    label = _clean_detail_text(label)
    label = re.sub(r"\s*[:：]\s*$", "", label)
    return label.strip()


def _clean_detail_value(value: Any, label: str) -> str:
    value_text = _clean_detail_text(value)
    label_text = _normalize_detail_label(label)
    if not value_text:
        return ""

    if label_text:
        escaped = re.escape(label_text)
        value_text = re.sub(rf"^{escaped}\s*[:：]?\s*", "", value_text, flags=re.I)
        value_text = re.sub(rf"\s*[:：]\s*{escaped}\s*[:：]?\s*$", "", value_text, flags=re.I)
        value_text = re.sub(rf"\s*\|\s*{escaped}\s*[:：]?\s*$", "", value_text, flags=re.I)

    return value_text.strip(" :-|")


def _merge_unique(existing: Any, new_value: Any) -> str:
    existing_text = _clean_detail_text(existing)
    new_text = _clean_detail_text(new_value)
    if not new_text:
        return existing_text
    if not existing_text:
        return new_text

    parts = [part.strip() for part in existing_text.split(" | ") if part.strip()]
    if new_text not in parts:
        parts.append(new_text)
    return " | ".join(parts)


def _add_product_detail(data: dict[str, Any], details_parts: list[str], label: Any, value: Any) -> None:
    label_text = _normalize_detail_label(label)
    value_text = _clean_detail_value(value, label_text)
    if not label_text or not value_text or label_text.lower().startswith("customer"):
        return

    details_parts.append(f"{label_text}: {value_text}")
    column = DETAIL_LABEL_TO_COLUMN.get(label_text)
    if column and column != "asin":
        data[column] = _merge_unique(data.get(column, ""), value_text)


def _parse_product_details_text(raw_details: Any) -> dict[str, str]:
    """Split existing raw product_details text into known detail columns."""
    text = _clean_detail_text(raw_details)
    if not text:
        return {}

    labels = sorted(DETAIL_LABEL_TO_COLUMN, key=len, reverse=True)
    label_pattern = "|".join(re.escape(label) for label in labels)
    pattern = re.compile(rf"(?:^|\s+\|\s+)({label_pattern})\s*[:：]\s*", flags=re.I)
    matches = list(pattern.finditer(text))
    details: dict[str, str] = {}

    for index, match in enumerate(matches):
        raw_label = match.group(1)
        label = next((candidate for candidate in labels if candidate.lower() == raw_label.lower()), raw_label)
        column = DETAIL_LABEL_TO_COLUMN.get(label)
        if not column or column == "asin":
            continue

        value_end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        value = _clean_detail_value(text[match.end():value_end], label)
        if value:
            details[column] = _merge_unique(details.get(column, ""), value)

    return details


def _order_export_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    for column in EXPORT_COLUMNS:
        if column not in df.columns:
            df[column] = ""

    leading_columns = [column for column in EXPORT_COLUMNS if column != "product_details"]
    extra_columns = [
        column for column in df.columns
        if column not in EXPORT_COLUMNS and column != "product_details"
    ]
    return df[leading_columns + extra_columns + ["product_details"]]


def _write_products_excel(df: pd.DataFrame, filename: Path) -> None:
    df = _order_export_dataframe(df.copy())
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Products")
        ws = writer.sheets["Products"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for column_index, column_name in enumerate(df.columns, start=1):
            letter = get_column_letter(column_index)
            if column_name in {"description", "features", "images", "product_details"}:
                width = 60
            elif column_name in {"url", "title"}:
                width = 45
            else:
                sample = df[column_name].fillna("").astype(str).head(100)
                max_len = max([len(str(column_name)), *(len(value) for value in sample)])
                width = min(max(max_len + 2, 12), 35)
            ws.column_dimensions[letter].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def organize_existing_excel(source_path: str | Path, output_path: str | Path | None = None) -> Path:
    """Create an organized copy of an existing scraper export."""
    source = Path(source_path)
    output = Path(output_path) if output_path else source.with_name(f"{source.stem}_organized{source.suffix}")
    df = pd.read_excel(source)

    if "product_details" in df.columns:
        for index, raw_details in df["product_details"].items():
            parsed_details = _parse_product_details_text(raw_details)
            for column, value in parsed_details.items():
                if column not in df.columns:
                    df[column] = ""
                df.at[index, column] = _merge_unique(df.at[index, column], value)

    _write_products_excel(df, output)
    return output


def _parse_amazon_page(html: str, url: str, asin: str) -> dict[str, Any]:
    """Extract product data from an Amazon product detail page."""
    soup = BeautifulSoup(html, "lxml")
    data: dict[str, Any] = {
        "asin": asin,
        "url": url,
        "title": "",
        "brand": "",
        "current_price": "",
        "original_price": "",
        "discount": "",
        "rating": "",
        "rating_count": "",
        "review_count": "",
        "description": "",
        "category": "",
        "images": "",
        "video_url": "",
        "colors": "",
        "sizes": "",
        "material": "",
        "seller": "",
        "availability": "",
        "features": "",
    }
    for column in DETAIL_COLUMNS:
        data[column] = ""
    data["product_details"] = ""

    # ================================================================
    # 1. Title
    # ================================================================
    title_el = soup.select_one("#productTitle")
    if title_el:
        data["title"] = _clean_text(title_el.get_text())
    
    # Fallback title
    if not data["title"]:
        title_el = soup.select_one("h1 span.a-text-normal, h1.a-size-large")
        if title_el:
            data["title"] = _clean_text(title_el.get_text())

    # ================================================================
    # 2. Brand
    # ================================================================
    brand_el = soup.select_one("#bylineInfo, .po-brand .a-span9 span")
    if brand_el:
        brand_text = _clean_text(brand_el.get_text())
        # Clean "Visit the X Store" or "Brand: X"
        brand_text = re.sub(r'^Visit the\s+', '', brand_text)
        brand_text = re.sub(r'\s+Store$', '', brand_text)
        brand_text = re.sub(r'^Brand:\s*', '', brand_text)
        data["brand"] = brand_text
    
    # Try alternate brand location
    if not data["brand"]:
        brand_row = soup.select_one("tr.po-brand td.a-span9 span")
        if brand_row:
            data["brand"] = _clean_text(brand_row.get_text())

    # ================================================================
    # 3. Prices
    # ================================================================
    # Current price - multiple possible selectors
    price_selectors = [
        ".priceToPay span.a-price-whole",
        "#corePrice_feature_div .a-price-whole",
        "#priceblock_ourprice",
        "#priceblock_dealprice",
        "#priceblock_saleprice",
        ".a-price.a-text-price.a-size-medium span.a-offscreen",
        "#price_inside_buybox",
        "#newBuyBoxPrice",
        ".apexPriceToPay span.a-offscreen",
        "#apex_offerDisplay_desktop .a-price span.a-offscreen",
    ]
    
    for sel in price_selectors:
        price_el = soup.select_one(sel)
        if price_el:
            price_text = _clean_text(price_el.get_text())
            # Extract price value
            price_match = re.search(r'[\d,]+\.?\d*', price_text.replace(',', ''))
            if price_match:
                data["current_price"] = price_match.group(0)
                break
    
    # Original/MRP price
    mrp_selectors = [
        ".basisPrice span.a-offscreen",
        "#listPrice",
        ".a-text-strike",
        ".priceBlockStrikePriceString",
        "span.a-price.a-text-price span.a-offscreen",
    ]
    
    for sel in mrp_selectors:
        mrp_el = soup.select_one(sel)
        if mrp_el:
            mrp_text = _clean_text(mrp_el.get_text())
            mrp_match = re.search(r'[\d,]+\.?\d*', mrp_text.replace(',', ''))
            if mrp_match and mrp_match.group(0) != data["current_price"]:
                data["original_price"] = mrp_match.group(0)
                break
    
    # Discount percentage
    discount_el = soup.select_one(".savingsPercentage, #dealprice_savings .priceBlockSavingsString, .reinventPriceSavingsPercentageMargin")
    if discount_el:
        data["discount"] = _clean_text(discount_el.get_text())
    
    # Calculate discount if not found
    if not data["discount"] and data["current_price"] and data["original_price"]:
        try:
            cp = float(data["current_price"].replace(",", ""))
            op = float(data["original_price"].replace(",", ""))
            if op > cp:
                pct = round((1 - cp / op) * 100)
                data["discount"] = f"-{pct}%"
        except (ValueError, ZeroDivisionError):
            pass

    # ================================================================
    # 4. Ratings
    # ================================================================
    rating_el = soup.select_one("#acrPopover span.a-icon-alt, .reviewCountTextLinkedHistogram span.a-icon-alt")
    if rating_el:
        rating_text = rating_el.get_text()
        rating_match = re.search(r'([\d.]+)\s*out\s*of\s*5', rating_text)
        if rating_match:
            data["rating"] = rating_match.group(1)
    
    # Rating count
    rating_count_el = soup.select_one("#acrCustomerReviewText")
    if rating_count_el:
        count_text = rating_count_el.get_text()
        count_match = re.search(r'([\d,]+)', count_text)
        if count_match:
            data["rating_count"] = count_match.group(1).replace(",", "")
            data["review_count"] = data["rating_count"]

    # ================================================================
    # 5. Description / Features
    # ================================================================
    # Feature bullets
    feature_bullets = soup.select("#feature-bullets li span.a-list-item")
    features_list = []
    for bullet in feature_bullets:
        text = _clean_text(bullet.get_text())
        if text and not text.startswith("Make sure"):
            features_list.append(text)
    if features_list:
        data["features"] = " | ".join(features_list[:10])
    
    # Product description
    desc_el = soup.select_one("#productDescription p, #productDescription")
    if desc_el:
        data["description"] = _clean_text(desc_el.get_text())[:1000]
    
    # A+ Content description fallback
    if not data["description"]:
        aplus = soup.select_one("#aplus .aplus-v2")
        if aplus:
            data["description"] = _clean_text(aplus.get_text())[:1000]

    # ================================================================
    # 6. Category / Breadcrumbs
    # ================================================================
    breadcrumbs = soup.select("#wayfinding-breadcrumbs_feature_div li a")
    if breadcrumbs:
        categories = [_clean_text(bc.get_text()) for bc in breadcrumbs]
        data["category"] = " > ".join(cat for cat in categories if cat)

    # ================================================================
    # 7. Images
    # ================================================================
    images_list = []
    
    # Method 1: Look for image data in scripts
    for script in soup.find_all("script"):
        txt = script.string or ""
        if "'colorImages'" in txt or '"colorImages"' in txt or "'initial'" in txt:
            # Extract image URLs from the colorImages object
            img_urls = re.findall(r'"hiRes"\s*:\s*"([^"]+)"', txt)
            if not img_urls:
                img_urls = re.findall(r'"large"\s*:\s*"([^"]+)"', txt)
            for url in img_urls:
                if url and url.startswith("http") and url not in images_list:
                    images_list.append(url)
    
    # Method 2: Image gallery elements
    if not images_list:
        img_els = soup.select("#altImages img, #imageBlock img, #imgTagWrapperId img")
        for img in img_els:
            src = img.get("src") or img.get("data-old-hires") or img.get("data-a-dynamic-image")
            if src and "sprite" not in src and "blank" not in src:
                # Convert thumbnail to large image
                large_src = re.sub(r'\._[A-Z]+\d+_\.', '.', src)
                if large_src not in images_list:
                    images_list.append(large_src)
    
    # Method 3: Main image
    if not images_list:
        main_img = soup.select_one("#landingImage, #imgBlkFront")
        if main_img:
            src = main_img.get("data-old-hires") or main_img.get("src")
            if src:
                images_list.append(src)
    
    if images_list:
        data["images"] = " | ".join(images_list[:10])

    # ================================================================
    # 8. Video URL
    # ================================================================
    for script in soup.find_all("script"):
        txt = script.string or ""
        video_url = re.search(r'"url"\s*:\s*"([^"]+\.mp4[^"]*)"', txt)
        if video_url:
            data["video_url"] = video_url.group(1).replace("\\u002F", "/")
            break

    # ================================================================
    # 9. Colors / Variations
    # ================================================================
    color_buttons = soup.select("#variation_color_name li img, #color_name_0 option")
    colors = []
    for el in color_buttons:
        color = el.get("alt") or el.get_text()
        if color and color.strip() and color not in colors:
            colors.append(_clean_text(color))
    if colors:
        data["colors"] = ", ".join(colors[:20])
    
    # Also check for color in title or product details
    if not data["colors"]:
        color_row = soup.select_one("tr.po-color td.a-span9 span")
        if color_row:
            data["colors"] = _clean_text(color_row.get_text())

    # ================================================================
    # 10. Sizes
    # ================================================================
    size_buttons = soup.select("#variation_size_name option, #native_dropdown_selected_size_name option")
    sizes = []
    for el in size_buttons:
        size = _clean_text(el.get_text())
        if size and size.lower() != "select" and size not in sizes:
            sizes.append(size)
    if sizes:
        data["sizes"] = ", ".join(sizes[:30])
    
    # Size from product details
    if not data["sizes"]:
        size_row = soup.select_one("tr.po-size td.a-span9 span")
        if size_row:
            data["sizes"] = _clean_text(size_row.get_text())

    # ================================================================
    # 11. Material
    # ================================================================
    material_selectors = [
        "tr.po-material td.a-span9 span",
        "#productDetails_techSpec_section_1 tr:-soup-contains('Material') td",
    ]
    for sel in material_selectors:
        material_el = soup.select_one(sel)
        if material_el:
            data["material"] = _clean_text(material_el.get_text())
            break
    
    # Search in product details table
    if not data["material"]:
        for row in soup.select("#productDetails_techSpec_section_1 tr, #productDetails_detailBullets_sections1 tr"):
            cells = row.select("th, td")
            if len(cells) >= 2:
                key = _clean_text(cells[0].get_text()).lower()
                if "material" in key or "fabric" in key:
                    data["material"] = _clean_text(cells[1].get_text())
                    break

    # ================================================================
    # 12. Seller
    # ================================================================
    seller_el = soup.select_one("#sellerProfileTriggerId, #merchant-info a")
    if seller_el:
        data["seller"] = _clean_text(seller_el.get_text())
    
    # Sold by Amazon check
    if not data["seller"]:
        merchant_info = soup.select_one("#merchant-info")
        if merchant_info:
            text = _clean_text(merchant_info.get_text())
            if "Amazon" in text:
                data["seller"] = "Amazon"

    # ================================================================
    # 13. Availability
    # ================================================================
    avail_el = soup.select_one("#availability span, #outOfStock span")
    if avail_el:
        data["availability"] = _clean_text(avail_el.get_text())

    # ================================================================
    # 14. Product Details Table
    # ================================================================
    details_parts = []
    for row in soup.select("#productDetails_techSpec_section_1 tr, #productDetails_detailBullets_sections1 tr"):
        key_el = row.select_one("th")
        value_el = row.select_one("td")
        if key_el and value_el:
            _add_product_detail(data, details_parts, key_el.get_text(), value_el.get_text())

    for item in soup.select("#detailBullets_feature_div li"):
        key_el = item.select_one("span.a-text-bold")
        if key_el:
            _add_product_detail(data, details_parts, key_el.get_text(), item.get_text(" ", strip=True))

    if details_parts:
        data["product_details"] = " | ".join(details_parts)

    return data


# ---------------------------------------------------------------------------
# Core async scraping
# ---------------------------------------------------------------------------

async def _fetch(client: httpx.AsyncClient, url: str) -> str | None:
    """GET a URL with retries."""
    for attempt in range(3):
        try:
            headers = get_headers()
            resp = await client.get(url, headers=headers, follow_redirects=True, timeout=25)
            if resp.status_code == 200:
                return resp.text
            elif resp.status_code == 503:
                _log(f"⚠ Amazon rate limit (503), waiting...")
                await asyncio.sleep(random.uniform(5.0, 10.0))
            else:
                _log(f"⚠ HTTP {resp.status_code} for {url}")
        except httpx.HTTPError as exc:
            _log(f"⚠ Attempt {attempt+1} error: {exc}")
        await asyncio.sleep(random.uniform(2.0, 4.0))
    return None


async def _scrape_by_asin(
    sem: asyncio.Semaphore,
    client: httpx.AsyncClient,
    asin: str,
    is_retry: bool = False,
) -> dict[str, Any] | str | None:
    """Scrape a single product by its Amazon ASIN.
    
    Returns:
        dict: Product data if successful
        str: "CAPTCHA" if blocked by CAPTCHA (for retry)
        None: If failed for other reasons
    """
    async with sem:
        if scraper_state["stop_event"].is_set():
            return None
        # Random delay between requests to avoid detection
        # Longer delay for retries
        if is_retry:
            await asyncio.sleep(random.uniform(5.0, 10.0))
        else:
            await asyncio.sleep(random.uniform(1.5, 3.5))

        # Use Amazon India URL (can be changed to .com)
        url = f"https://www.amazon.in/dp/{asin}"
        html = await _fetch(client, url)

        if not html:
            _log(f"✗ Failed to fetch ASIN: {asin}")
            scraper_state["failed_count"] += 1
            scraper_state["scraped_count"] += 1
            count = scraper_state["scraped_count"]
            total = scraper_state["total_items"]
            _log(f"→ [{count}/{total}] FAILED — {asin}")
            return None

        # Check for CAPTCHA
        if "Enter the characters you see below" in html or "api-services-support@amazon" in html:
            if is_retry:
                # Already a retry, count as failed
                _log(f"⚠ CAPTCHA again on retry for ASIN: {asin}")
                scraper_state["failed_count"] += 1
                scraper_state["scraped_count"] += 1
                count = scraper_state["scraped_count"]
                total = scraper_state["total_items"]
                _log(f"→ [{count}/{total}] CAPTCHA BLOCKED (retry failed) — {asin}")
                return None
            else:
                # First attempt, mark for retry later
                _log(f"⚠ CAPTCHA detected for ASIN: {asin} (will retry later)")
                return "CAPTCHA:" + asin

        data = _parse_amazon_page(html, url, asin)
        scraper_state["scraped_count"] += 1
        count = scraper_state["scraped_count"]
        total = scraper_state["total_items"]
        label = data.get("title")[:50] + "..." if data.get("title") and len(data.get("title", "")) > 50 else data.get("title") or asin
        _log(f"→ [{count}/{total}] {label}")
        return data


async def scrape_task(asins: list[str]) -> None:
    """Main background scraping coroutine."""
    state = scraper_state
    state["running"] = True
    state["stop_event"].clear()
    state["logs"].clear()
    state["products"].clear()
    state["scraped_count"] = 0
    state["failed_count"] = 0
    state["captcha_blocked"] = []
    state["total_items"] = len(asins)
    state["output_file"] = None

    _log(f"🔍 Starting scrape for {len(asins)} ASINs…")

    captcha_asins = []  # Track ASINs blocked by CAPTCHA

    async with httpx.AsyncClient(http2=False) as client:
        sem = asyncio.Semaphore(MAX_CONCURRENCY)

        # First pass: scrape all ASINs
        tasks = []
        for asin in asins:
            if state["stop_event"].is_set():
                break
            tasks.append(_scrape_by_asin(sem, client, str(asin), is_retry=False))

        results = await asyncio.gather(*tasks, return_exceptions=True)

        for r in results:
            if isinstance(r, dict):
                state["products"].append(r)
            elif isinstance(r, str) and r.startswith("CAPTCHA:"):
                # Extract ASIN from CAPTCHA marker
                captcha_asin = r.replace("CAPTCHA:", "")
                captcha_asins.append(captcha_asin)

        # Retry CAPTCHA-blocked ASINs if any
        if captcha_asins and not state["stop_event"].is_set():
            _log(f"")
            _log(f"🔄 Retrying {len(captcha_asins)} CAPTCHA-blocked ASINs after delay…")
            _log(f"⏳ Waiting 30 seconds before retry to avoid detection…")
            await asyncio.sleep(30)  # Wait before retrying
            
            retry_tasks = []
            for asin in captcha_asins:
                if state["stop_event"].is_set():
                    break
                retry_tasks.append(_scrape_by_asin(sem, client, asin, is_retry=True))
            
            retry_results = await asyncio.gather(*retry_tasks, return_exceptions=True)
            
            for r in retry_results:
                if isinstance(r, dict):
                    state["products"].append(r)
                    _log(f"✅ Retry successful for: {r.get('asin', 'unknown')}")

    # Save Excel
    if state["products"]:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = DOWNLOADS_DIR / f"amazon_asin_{ts}.xlsx"
        df = pd.DataFrame(state["products"])
        _write_products_excel(df, filename)
        state["output_file"] = str(filename)
        _log(f"✅ Saved {len(state['products'])} products → {filename.name}")
        if state["failed_count"] > 0:
            _log(f"⚠ {state['failed_count']} ASINs failed to scrape.")
    else:
        _log("⚠ No products scraped — nothing to save.")

    if state["stop_event"].is_set():
        _log("🛑 STOPPED by user.")
    else:
        _log("🏁 DONE")

    state["running"] = False


# ---------------------------------------------------------------------------
# FastAPI routes
# ---------------------------------------------------------------------------

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(request, "index.html")


@app.post("/upload")
async def upload_and_start(file: UploadFile = File(...)):
    """Upload an Excel file with ASINs and start scraping."""
    if scraper_state["running"]:
        return {"status": "error", "message": "A scraping task is already running."}

    # Validate file type
    fname = file.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls")):
        return {"status": "error", "message": "Please upload an .xlsx or .xls file."}

    # Save uploaded file
    save_path = UPLOADS_DIR / fname
    contents = await file.read()
    with open(save_path, "wb") as f:
        f.write(contents)

    # Read ASINs
    try:
        df = pd.read_excel(save_path)
    except Exception as e:
        return {"status": "error", "message": f"Could not read Excel file: {e}"}

    # Find the ASIN column (case-insensitive)
    asin_col = None
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if col_lower in ("asin", "asins", "asin_id", "product_id", "productid"):
            asin_col = col
            break
        if "asin" in col_lower:
            asin_col = col
            break
    if asin_col is None:
        # Use first column as fallback
        asin_col = df.columns[0]

    asins = df[asin_col].dropna().astype(str).str.strip().tolist()
    # Clean ASINs
    asins = [asin.replace(".0", "") for asin in asins if asin and asin != "nan"]
    # Validate ASIN format (10 alphanumeric characters)
    valid_asins = [asin for asin in asins if re.match(r'^[A-Z0-9]{10}$', asin.upper())]
    
    if not valid_asins:
        return {"status": "error", "message": "No valid ASINs found in the uploaded file. ASINs should be 10 alphanumeric characters."}

    if IS_VERCEL:
        await scrape_task([a.upper() for a in valid_asins])
        filepath = scraper_state.get("output_file")
        if not filepath or not os.path.isfile(filepath):
            return {
                "status": "error",
                "message": "Scraping finished, but no Excel file was generated.",
                "logs": scraper_state["logs"],
            }

        with open(filepath, "rb") as f:
            file_base64 = base64.b64encode(f.read()).decode("ascii")

        return {
            "status": "completed",
            "message": "Scraping completed. Excel file is ready.",
            "total": len(valid_asins),
            "scraped": scraper_state["scraped_count"],
            "failed": scraper_state["failed_count"],
            "logs": scraper_state["logs"],
            "file_name": os.path.basename(filepath),
            "file_base64": file_base64,
        }

    # Launch background task
    loop = asyncio.get_event_loop()
    scraper_state["task"] = loop.create_task(scrape_task([a.upper() for a in valid_asins]))

    return {
        "status": "started",
        "message": f"Found {len(valid_asins)} ASINs. Scraping started!",
        "total": len(valid_asins),
    }


@app.post("/stop")
async def stop_scraping():
    if not scraper_state["running"]:
        return {"status": "info", "message": "No active task."}
    scraper_state["stop_event"].set()
    return {"status": "stopped", "message": "Stop signal sent."}


@app.get("/progress")
async def progress():
    """SSE endpoint streaming log lines."""
    async def event_generator():
        last_index = 0
        while True:
            logs = scraper_state["logs"]
            while last_index < len(logs):
                yield f"data: {logs[last_index]}\n\n"
                last_index += 1

            if not scraper_state["running"] and last_index >= len(logs):
                if scraper_state.get("output_file"):
                    yield f"data: __FILE_READY__\n\n"
                yield f"data: __END__\n\n"
                break

            await asyncio.sleep(0.4)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/download")
async def download():
    filepath = scraper_state.get("output_file")
    if not filepath or not os.path.isfile(filepath):
        return {"status": "error", "message": "No file available for download."}
    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(filepath),
    )


@app.get("/status")
async def status():
    return {
        "running": scraper_state["running"],
        "scraped": scraper_state["scraped_count"],
        "total": scraper_state["total_items"],
        "failed": scraper_state["failed_count"],
        "has_file": scraper_state.get("output_file") is not None,
    }


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8002))
    uvicorn.run("app:app", host="0.0.0.0", port=port, reload=True)
